"""Export scored email data to Excel."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.email import Email
from app.models.score import Score

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

HEADER_FONT = Font(name="Arial", bold=True)
BODY_FONT = Font(name="Arial")

SCORE_DIMS = ["personalisation", "clarity", "value_proposition", "cta", "overall"]

EMAIL_SCORES_HEADERS = [
    "Rep", "Subject", "Date", "Personalisation", "Clarity",
    "Value Proposition", "CTA", "Overall", "Notes",
]

REP_AVERAGES_HEADERS = [
    "Rep", "Personalisation", "Clarity", "Value Proposition", "CTA", "Overall",
]


def _score_fill(value: int | None) -> PatternFill | None:
    """Return the fill colour for a score value."""
    if value is None:
        return None
    if value >= 8:
        return GREEN_FILL
    if value >= 6:
        return YELLOW_FILL
    if value >= 4:
        return ORANGE_FILL
    return RED_FILL


async def export_to_excel(session: AsyncSession, output_path: str) -> str:
    """Export scored emails and rep averages to an xlsx workbook.

    Sheet 1 "Email Scores": one row per scored email with colour-coded score cells.
    Sheet 2 "Rep Averages": one row per rep, sorted by overall average descending.

    Returns the output_path.
    """
    wb = Workbook()

    # --- Sheet 1: Email Scores ---
    ws = wb.active
    ws.title = "Email Scores"
    ws.append(EMAIL_SCORES_HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"

    stmt = (
        select(Email, Score)
        .join(Score, Email.id == Score.email_id)
        .where(Score.score_error.is_(False))
    )
    result = await session.execute(stmt)
    rows = result.all()

    for email, score in rows:
        ws.append([
            email.from_name or email.from_email,
            email.subject,
            email.timestamp,
            score.personalisation,
            score.clarity,
            score.value_proposition,
            score.cta,
            score.overall,
            score.notes,
        ])
        row_num = ws.max_row
        for col in range(1, len(EMAIL_SCORES_HEADERS) + 1):
            ws.cell(row=row_num, column=col).font = BODY_FONT
        # Colour-code score columns (4 through 8)
        for i, dim in enumerate(SCORE_DIMS):
            cell = ws.cell(row=row_num, column=4 + i)
            fill = _score_fill(getattr(score, dim))
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions

    # --- Sheet 2: Rep Averages ---
    ws2 = wb.create_sheet("Rep Averages")
    ws2.append(REP_AVERAGES_HEADERS)
    for cell in ws2[1]:
        cell.font = HEADER_FONT
    ws2.freeze_panes = "A2"

    avg_stmt = (
        select(
            Email.from_email,
            func.avg(Score.personalisation).label("avg_personalisation"),
            func.avg(Score.clarity).label("avg_clarity"),
            func.avg(Score.value_proposition).label("avg_value_proposition"),
            func.avg(Score.cta).label("avg_cta"),
            func.avg(Score.overall).label("avg_overall"),
        )
        .join(Score, Email.id == Score.email_id)
        .where(Score.score_error.is_(False))
        .group_by(Email.from_email)
        .order_by(func.avg(Score.overall).desc())
    )
    result = await session.execute(avg_stmt)

    for row in result.all():
        ws2.append([
            row.from_email,
            round(float(row.avg_personalisation), 1),
            round(float(row.avg_clarity), 1),
            round(float(row.avg_value_proposition), 1),
            round(float(row.avg_cta), 1),
            round(float(row.avg_overall), 1),
        ])
        row_num = ws2.max_row
        for col in range(1, len(REP_AVERAGES_HEADERS) + 1):
            ws2.cell(row=row_num, column=col).font = BODY_FONT

    ws2.auto_filter.ref = ws2.dimensions

    wb.save(output_path)
    return output_path

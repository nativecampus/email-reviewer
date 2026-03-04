from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.services.export import export_to_excel


class TestExportToExcel:
    async def test_creates_file_at_given_path(
        self, db, make_email, make_score, tmp_path
    ):
        email = await make_email()
        await make_score(email_id=email.id)

        output_path = str(tmp_path / "report.xlsx")
        result = await export_to_excel(db, output_path)

        assert (tmp_path / "report.xlsx").exists()
        assert result == output_path

    async def test_email_scores_sheet_has_expected_headers(
        self, db, make_email, make_score, tmp_path
    ):
        email = await make_email()
        await make_score(email_id=email.id)

        output_path = str(tmp_path / "report.xlsx")
        await export_to_excel(db, output_path)

        wb = load_workbook(output_path)
        assert "Email Scores" in wb.sheetnames
        ws = wb["Email Scores"]
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "Rep",
            "Subject",
            "Date",
            "Personalisation",
            "Clarity",
            "Value Proposition",
            "CTA",
            "Overall",
            "Notes",
        ]

    async def test_rep_averages_sheet_has_expected_headers(
        self, db, make_email, make_score, tmp_path
    ):
        email = await make_email()
        await make_score(email_id=email.id)

        output_path = str(tmp_path / "report.xlsx")
        await export_to_excel(db, output_path)

        wb = load_workbook(output_path)
        assert "Rep Averages" in wb.sheetnames
        ws = wb["Rep Averages"]
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "Rep",
            "Personalisation",
            "Clarity",
            "Value Proposition",
            "CTA",
            "Overall",
        ]

    async def test_score_colour_coding(
        self, db, make_email, make_score, tmp_path
    ):
        email = await make_email()
        await make_score(
            email_id=email.id,
            personalisation=8,
            clarity=6,
            value_proposition=4,
            cta=2,
            overall=5,
        )

        output_path = str(tmp_path / "report.xlsx")
        await export_to_excel(db, output_path)

        wb = load_workbook(output_path)
        ws = wb["Email Scores"]

        green = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        yellow = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )
        orange = PatternFill(
            start_color="F4B084", end_color="F4B084", fill_type="solid"
        )
        red = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        # Row 2 data; cols 4-7 are Personalisation, Clarity, Value Proposition, CTA
        assert ws.cell(row=2, column=4).fill == green   # score 8 >= 8
        assert ws.cell(row=2, column=5).fill == yellow  # score 6 >= 6
        assert ws.cell(row=2, column=6).fill == orange  # score 4 >= 4
        assert ws.cell(row=2, column=7).fill == red     # score 2 < 4

    async def test_rep_averages_sorted_by_overall_descending(
        self, db, make_email, make_score, tmp_path
    ):
        e1 = await make_email(from_email="high@example.com")
        await make_score(
            email_id=e1.id,
            personalisation=9, clarity=9, value_proposition=9, cta=9, overall=9,
        )

        e2 = await make_email(from_email="mid@example.com")
        await make_score(
            email_id=e2.id,
            personalisation=6, clarity=6, value_proposition=6, cta=6, overall=6,
        )

        e3 = await make_email(from_email="low@example.com")
        await make_score(
            email_id=e3.id,
            personalisation=3, clarity=3, value_proposition=3, cta=3, overall=3,
        )

        output_path = str(tmp_path / "report.xlsx")
        await export_to_excel(db, output_path)

        wb = load_workbook(output_path)
        ws = wb["Rep Averages"]

        reps = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert reps == ["high@example.com", "mid@example.com", "low@example.com"]
